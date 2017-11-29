from openpyxl import load_workbook
import xlrd
import csv
import sys
import random
import os

GroupColumnIndex = 1
GROUP_VALUES = {"low":0, "high":1,"clinical":1}
class DataFilesObj:
    xml_path = ""
    sheet = ""
    data_list = []
    csv_files_dir = None
    titles = []
    def __init__(self,xml_path="OrganizedData.xlsx", sheet="Sheet1"):
        self.xml_path = xml_path
        self.sheet = sheet

    def choose_features_by_column_num (self, columns_index_list, start_counting_column):
        workbook = xlrd.open_workbook(self.xml_path, on_demand = True)
        sheet = workbook.sheet_by_name(self.sheet)
        for rownum in range(sheet.nrows):
            single_subject_data_list = []
            single_subject_data_list.append(sheet.cell(rownum, GroupColumnIndex).value)
            for index in columns_index_list:
                single_subject_data_list.append(sheet.cell(rownum,index+start_counting_column).value)
            self.data_list.append(single_subject_data_list)
        self.organize_data()


    def choose_features_by_title (self, titles_list):
        workbook = xlrd.open_workbook(self.xml_path, on_demand = True)
        sheet = workbook.sheet_by_name(self.sheet)
        columns_index_list= []
        for colnum in range(sheet.ncols):
            column_tile = sheet.cell(0, colnum)
            if column_tile in titles_list:
                columns_index_list.append(colnum)
        for rownum in range(sheet.nrows):
            single_subject_data_list = []
            single_subject_data_list.append(sheet.cell(rownum, GroupColumnIndex).value)
            for index in columns_index_list:
                single_subject_data_list.append(sheet.cell(index).value)
        self.data_list.append(single_subject_data_list)
        self.organize_data()

    def organize_data (self):
        self.titles = self.data_list[0]
        self.data_list.remove(self.data_list[0])
        for single_subject_data in self.data_list:
            single_subject_data[0]=GROUP_VALUES.get(single_subject_data[0])

    def split_data (self, num_of_tests):
        random.shuffle (self.data_list)
        test_list = self.data_list[:num_of_tests]
        training_list = self.data_list[num_of_tests:len(self.data_list)]
        return training_list,test_list

    def create_csv_files (self,num_of_tests,dir):
        if num_of_tests > len(self.data_list)-1 or num_of_tests<1:
            return
        training_data,test_data = self.split_data(num_of_tests)

        if not os.path.exists(dir):
            os.makedirs(dir)


        f = open(os.path.join(dir,"training_data.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow(self.titles[1:len(self.titles)])
        for i in range(len(training_data)):
            row = []
            for j in range(len(training_data[0])-1):
                row.append(training_data[i][j+1])
            writer.writerow(row)
        f.close()


        f = open(os.path.join(dir,"test_data.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow(self.titles[1:len(self.titles)])
        for i in range(len(test_data)):
            row = []
            for j in range(len(test_data[0])-1):
                row.append(test_data[i][j+1])
            writer.writerow(row)
        f.close()

        f = open(os.path.join(dir,"training_labels.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow([self.titles[0]])
        for i in range(len(training_data)):
            writer.writerow([training_data[i][0]])
        f.close()

        f = open(os.path.join(dir,"test_labels.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow([self.titles[0]])
        for i in range(len(test_data)):
            writer.writerow([test_data[i][0]])
        f.close()


    def create_csv_files_with_premade_data (self,training_data,test_data,dir):
        if not os.path.exists(dir):
            os.makedirs(dir)

        f = open(os.path.join(dir,"training_data.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow(self.titles[1:len(self.titles)])
        for i in range(len(training_data)):
            row = []
            for j in range(len(training_data[0])-1):
                row.append(training_data[i][j+1])
            writer.writerow(row)
        f.close()


        f = open(os.path.join(dir,"test_data.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow(self.titles[1:len(self.titles)])
        for i in range(len(test_data)):
            row = []
            for j in range(len(test_data[0])-1):
                row.append(test_data[i][j+1])
            writer.writerow(row)
        f.close()

        f = open(os.path.join(dir,"training_labels.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow([self.titles[0]])
        for i in range(len(training_data)):
            writer.writerow([training_data[i][0]])
        f.close()

        f = open(os.path.join(dir,"test_labels.csv"), 'wt')
        writer = csv.writer(f)
        writer.writerow([self.titles[0]])
        for i in range(len(test_data)):
            writer.writerow([test_data[i][0]])
        f.close()

